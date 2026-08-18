from __future__ import annotations

from pathlib import Path
from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from hockey_bot.models.enums import RU_LABELS
from hockey_bot.models.tables import Arena, Event, Expense, League, LeagueTeam, Season, Team
from hockey_bot.services.events import total_expenses
from hockey_bot.services.statistics import sports_stats


def export_excel(session: Session, user_id: int, path: str | Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "События"
    ws.append(["дата", "время", "тип события", "лига", "сезон", "собственная команда", "соперник", "дом/выезд", "арена", "адрес", "участие", "результат", "счёт", "стоимость", "расходы", "комментарий"])
    events = list(session.scalars(select(Event).where(Event.user_id == user_id).order_by(Event.event_date, Event.event_time)))
    for e in events:
        ws.append([e.event_date.isoformat(), e.event_time.strftime("%H:%M"), RU_LABELS[e.event_type], e.league_name_snapshot, e.season_name_snapshot, e.own_team_name_snapshot, e.opponent_name_snapshot, RU_LABELS.get(e.home_away, ""), e.arena_name_snapshot, e.arena_address_snapshot, RU_LABELS.get(e.participation, RU_LABELS.get(e.attendance, "")), RU_LABELS.get(e.result, ""), e.score, e.cost_rub, total_expenses(session, e.id), e.comment])
    ws = wb.create_sheet("Расходы")
    ws.append(["дата", "событие", "категория", "сумма"])
    for exp, event in session.execute(select(Expense, Event).join(Event, Expense.event_id == Event.id).where(Event.user_id == user_id)):
        ws.append([exp.created_at.date().isoformat(), event.id, RU_LABELS[event.event_type], exp.amount_rub])
    for title, model, headers in [("Лиги", League, ["название", "статус"]), ("Арены", Arena, ["название", "адрес", "статус"]), ("Сезоны", Season, ["название", "статус"]), ("Команды", Team, ["название", "статус", "принадлежность к лигам"] )]:
        ws = wb.create_sheet(title); ws.append(headers)
        for obj in session.scalars(select(model).order_by(model.id)):
            status = "Архив" if obj.archived else "Активна"
            if model is Arena: ws.append([obj.name, obj.address, status])
            elif model is Team:
                league_names = [session.get(League, lt.league_id).name for lt in session.scalars(select(LeagueTeam).where(LeagueTeam.team_id == obj.id))]
                ws.append([obj.name, status, ", ".join(league_names)])
            else: ws.append([obj.name, status])
    ws = wb.create_sheet("Статистика")
    ws.append(["показатель", "значение"])
    for key, value in sports_stats(session, user_id).items():
        ws.append([key, value])
    out = Path(path)
    wb.save(out)
    return out
